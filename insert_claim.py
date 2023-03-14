import openpyxl

wb = openpyxl.load_workbook(r'E:\!!Server\SSH.xlsx')

sheet = wb.active
# Sheet1
print(sheet)
print(sheet['R42'].value)
print(type(sheet['R42'].value))
print([sheet['R42'].value])
rows = sheet.max_row
cols = sheet.max_column
print('strings - ',rows)
print('columns - ',cols)
for i in range(2, rows + 1):
    db = openpyxl.Workbook()
    db.create_sheet(title='Main_inform', index=0)
    sheetdb = db['Main_inform']
    jdb = 1
    for j in range(2, cols+1):
        cell = sheet.cell(row=i, column=j)
        if (cell.value != None):
            resume = str(cell.value)
            celldb = sheetdb.cell(row=jdb, column=1)
            celldb.value = resume
            jdb +=1
            # print(resume)
    fname1 = 'E:/!!Server/Database_SSH/claims/'
    cellname = sheet.cell(row=i, column=2)
    fname2 = str(cellname.value)
    fname = fname1+fname2+'.SSH'
    db.save(fname)






    # E:\!!Server\Database_SSH\claims