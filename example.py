import openpyxl
from openpyxl.chart import *
# from openpyxl.chart import BarChart, Reference

# создаем новый excel-файл
wb = openpyxl.load_workbook('example.xlsx')
# добавляем новый лист
wb.create_sheet(title = 'Second лист', index = 1)
# получаем лист, с которым будем работать
sheet = wb['Second лист']

sheet['A1'] = 'Серия 1'
# это колонка с данными
for i in range(1, 11):
    cell = sheet.cell(row = i + 1, column = 1)
    cell.value = i * i

# создаем диаграмму
chart = BarChart()
chart.title = 'Первая серия данных'
data = Reference(sheet, min_col = 1, min_row = 1, max_col = 1, max_row = 11)
chart.add_data(data, titles_from_data = True)

# добавляем диаграмму на лист
sheet.add_chart(chart, 'C8')

# записываем файл
wb.save('example.xlsx')