import os
import openpyxl
import create_path
import tkinter as tk
import get_need_val
from tkinter.messagebox import showerror, showwarning, showinfo
import claim_create_descr



#######################################################################################################################
#                                   some function                                                                     #
#######################################################################################################################




#######################################################################################################################
#                                   Function to insert claim number                                                   #
#######################################################################################################################
def find_claim_number(link_all_cl, c_muser, c_yuser):
    claim_n = []
    claim_number = -1
    fullpath = create_path.createpath(link_all_cl, "claims_list.xlsx")

    if (os.path.exists(fullpath) == True): #Если уже есть файл с внесенными данными
        # Требуется исправить несовместимость строчного типа для месяца и года и числового при определении очередности
        wb = openpyxl.load_workbook(fullpath)
        sheet = wb['claims_list']
        rows = sheet.max_row
        for i in range(2, rows+1):
            if (str(sheet.cell(row = i, column = 4).value) ==  c_muser) & (str(sheet.cell(row = i, column = 5).value) ==  c_yuser):
                claim_n.append(int(sheet.cell(row = i, column = 3).value))
        if len(claim_n) > 0: #Присвоение номеру заявки №1, если он свободен
            claim_n.sort()
            if claim_n[0] > 1:
                claim_number = 1
            else:
                for j in range(1, len(claim_n)): # Определение первого свободного номера
                    if claim_n[j]-claim_n[j-1] > 1:
                        claim_number = claim_n[j-1]+1
                        break
            if claim_number == -1:
                claim_number = claim_n[-1]+1
        else:
            claim_number = 1
    else:                                           #Если это самая первая заявка, вносимая в базу данных
        if not os.path.isdir(link_all_cl):          #Проверка, есть ли такой путь, и если нет, то его создание
            os.makedirs(link_all_cl)

        wb = openpyxl.Workbook()                    #Если это самая первая заявка, вносимая в базу данных
        wb.create_sheet(title='claims_list', index=0)
        sheet = wb['claims_list']                   # шапка таблицы №	claim_letter	number_in_data	month	year	claim	claim_location
        sheet.cell(row=1, column=1).value = "№"
        sheet.cell(row=1, column=2).value = "claim_letter"
        sheet.cell(row=1, column=3).value = "number_in_data"
        sheet.cell(row=1, column=4).value = "month"
        sheet.cell(row=1, column=5).value = "year"
        sheet.cell(row=1, column=6).value = "claim"
        sheet.cell(row=1, column=7).value = "claim_location"
        wb.save(fullpath)
        claim_number = 1

    claim_number = str(claim_number)
    while len(claim_number) <4:
        claim_number = "0"+claim_number
    return claim_number

#######################################################################################################################
#                                   Function to check claim number                                                    #
#######################################################################################################################
def ch_claim(new_c_n, fullpath, c_muser, c_yuser):
    claim_n = []
    wb = openpyxl.load_workbook(fullpath)
    sheet = wb['claims_list']
    rows = sheet.max_row
    for i in range(2, rows + 1):
        if (str(sheet.cell(row=i, column=4).value) == c_muser) & (str(sheet.cell(row=i, column=5).value) == c_yuser):
            claim_n.append(int(sheet.cell(row=i, column=3).value))
    if int(new_c_n) in claim_n:
        return False
    else:
        return True

#######################################################################################################################
#                                   FOR TESTS                                                                         #
#######################################################################################################################
# import tkinter as tk
    # main_window2 = tk.Tk()
    # frame_but = tk.Frame(master=main_window2, bg="snow")
    # frame_but.pack(fill=tk.BOTH, side=tk.LEFT, expand=True)
    # lbl_region = tk.Label(master=main_window2, text=os.path.exists(fullpath), bg="snow", fg="black")  #
    # lbl_region.pack()
    # main_window2.mainloop()

    # claim_letter = family[0].upper()
    #
    #
    # window2 = tk.Toplevel()
    # window2.title("Проверка номера заявки")
    # frm_claimcorrect = tk.Frame(master=window2)
    # frm_claimcorrect.pack(fill=tk.X, ipadx=5, ipady=5)
    # claim_num_title = tk.Label(master=frm_claimcorrect, text="Заявлению присвоен следующий номер:  ", bg="snow", fg="black",
    #                        font=(font_type, font_size))
    # claim_num_title.grid(row=0, column=0, sticky="e")
    # claim_num_title_let = tk.Label(master=frm_claimcorrect, text=claim_letter, bg="snow",
    #                            fg="black",
    #                            font=(font_type, font_size))
    # claim_num_title_let.grid(row=0, column=1, sticky="e")
    #
    # claim_num_title_num = tk.Entry(master=frm_claimcorrect, width=4, font=(font_type, font_size))
    # claim_num_title_num.grid(row=0, column=2, sticky="e")
    # claim_num_title_num.insert(0, claim_number)
    #
    # claim_num_title_dat = tk.Label(master=frm_claimcorrect, text="-"+c_muser+"-"+c_yuser, bg="snow",
    #                                fg="black",
    #                                font=(font_type, font_size))
    # claim_num_title_dat.grid(row=0, column=3, sticky="e")
    #
    # btn_next2 = tk.Button(master=frm_claimcorrect, text="Далее", bg="snow", fg="black",
    #                      font=(font_type, font_size), command=check_claim)#!!!(claim_letter, claim_num_title_num.get(), c_muser, c_yuser))
    # btn_next2.grid(row=1, column=3, sticky="e")
    #
    # window2.mainloop()

    # claim_f = {
    #     # Основные данные
    #     'number': claim_number,
    #     'family': family,
    #     'name': name,
    #     'fathername': fathername,
    #     'iin': iin,
    #     'adresse': adresse,
    #     'region': region,
    #     # Дата подачи заявления
    #     'day': c_duser,
    #     'month': c_muser,
    #     'year': c_yuser,
    #     # Статус заявки
    #     'status': claim_status,
    #     # номер телефонов
    #     'tel_n': ph_nmb_cl,
    #     'tel_n_name': ph_nam_cl
    #        }